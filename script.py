#!/usr/bin/python
# -*- coding: UTF-8 -*-

import xml.sax
import sys 
import openpyxl
import os

from datetime import datetime,date,timedelta
from operator import itemgetter
from itertools import groupby 
 

AllRecords = list()
WeekDaysRecords = list()
Dateskey = list()


Author = ''
Days = 0
resultXml = ''

if len(sys.argv) > 1 :
	Author = sys.argv[1]
	Days = int(sys.argv[2])
	resultXml = sys.argv[3]




class LogRecords(xml.sax.ContentHandler):
	def __init__(self):
		self.record = dict()
		self.tag = ''
		self.content = ''	
		self.revision = ''	
		
	
	def startElement(self, tag, attrs):
		self.tag = tag
		self.content = ''
		if tag == 'logentry':
			self.revision = attrs.getValue('revision')
	
	def endElement(self, tag):
		if tag == 'logentry':
			self.record['revision'] = self.revision
			AllRecords.append(self.record.copy())
			self.record.clear()
			
			
			self.tag = ''
			self.content = ''
			self.revision = ''
		
		if self.tag != "":						
			self.content = 	self.content.strip('\n')
			if tag == 'date':
				self.content = self.content[0:10]
			self.record[self.tag] = self.content
			self.content = ''
			
	def characters(self, content):
		content = content.strip()
		if content != "":
			self.content += content
			self.content += "\n" 
	

	
	
def progressSummaryOfThisWeek(wb):
	print('本周计划总结')
	su = u'本周计划总结'
	ws = wb[su]
#	4 为 周5
#	0 为 周1
 	nowDate = datetime.now()
	summaryDate = nowDate.strftime("%Y.%m.%d")
	ws['M4'] = summaryDate
	
	for vdate_str in Dateskey:
		planDate = datetime.strptime(vdate_str, '%Y-%m-%d')
		if planDate.weekday() == 0:
			planDate = planDate.strftime("%Y.%m.%d")
			ws['F4'] = planDate
			
	print('本周计划总结ok..')
	
	
	
def progressDs(wb):
	print('进入日总结')
	su = u'日总结'
	ws = wb[su]
	
	nowDate = datetime.now()
	ws['G3'] = nowDate.strftime("%Y.%m.%d")
	lastRow = 6 + Days - 1
	i = 0
	for day in lstgall:
		if i == Days:
			break
		content = ''
		for record in day:
			content += record['msg']
			content += '\n'
			
		content = content.strip('\n')
		ws.cell(row= lastRow - i, column=3, value=content)
		i+=1
	
	print('日总结 处理ok..')
	

def progressNextWeek(wb):
	print('下周计划')
	
	su = u'下周计划'
	ws = wb[su]
	nowDate = datetime.now()
	ws['F3'] = nowDate.strftime("%Y.%m.%d")

	weekday = nowDate.weekday()
	
	if weekday < 4:
		nowDate = nowDate + timedelta(days=4-weekday)	
		
	
	nextWeek = nowDate + timedelta(days=7)
	mondayMonth = ''
	mondayDay = ''
	
	fridayMonth = ''
	fridayDay = ''
		
	for i in range(7):
		nextWeekDay = nextWeek - timedelta(days=i)
		if nextWeekDay.weekday() == 4:
			fridayMonth = str(nextWeekDay.month)
			fridayDay = str(nextWeekDay.day)
		if nextWeekDay.weekday() == 0:
			mondayMonth = str(nextWeekDay.month)
			mondayDay = str(nextWeekDay.day)
			break
	
	titleStr = str('下周计划（{}月{}日-{}月{}日）').format(mondayMonth, mondayDay,fridayMonth,fridayDay)
	
	ws['A2'] = titleStr
	
	print('下周计划 ok..')
	
	


if (__name__ == "__main__"):
	if resultXml == '':
		exit(0)
	
	print('处理Log数据中...')
	parser = xml.sax.make_parser()
	parser.setFeature(xml.sax.handler.feature_namespaces,0)
	handler = LogRecords()
	parser.setContentHandler(handler)
	parser.parse(resultXml)	
	data = list(filter(lambda x : x['author'] == Author, AllRecords))
	
	lstg = groupby(data,itemgetter('date')) 
	lstgall= [list(group) for key,group in lstg]
	
	
	dates = groupby(data,itemgetter('date')) 
	Dateskey= [key for key,group in dates]

	print('处理Log数据ok..')
#  xml 所有数据

	print('读取模版')
	parent_path = os.path.dirname(resultXml) 
	sample_path = parent_path + '/sample.xlsx'
	
	
	
	wb = openpyxl.load_workbook(sample_path)
	#	['附注', '本周计划总结', '日总结', '下周计划']
	#	[u'\u9644\u6ce8', u'\u672c\u5468\u8ba1\u5212\u603b\u7ed3', u'\u65e5\u603b\u7ed3', u'\u4e0b\u5468\u8ba1\u5212']
	
	progressDs(wb)
	progressSummaryOfThisWeek(wb)
	progressNextWeek(wb)
	
	nowDate = datetime.now()
	outFileName = str('/（智付技术部-旅游产品线）{} {}月{}日周计划总结.xlsx').format(Author,nowDate.month,nowDate.day)
	out_path = parent_path + outFileName
	
	wb.save(out_path)


